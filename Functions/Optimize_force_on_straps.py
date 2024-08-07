#!/usr/bin/env python
# coding: utf-8

# In[2]:


import scipy.optimize
from scipy.optimize import minimize
from scipy.optimize import Bounds
from scipy.optimize import LinearConstraint
import math
import numpy as np
from scipy.integrate import simps
from sympy import *
from sympy import symbols
import pandas as pd
import openpyxl
from openpyxl import Workbook


# In[ ]:


def force_on_straps(sum_Qy, sum_Mz, name_of_element, percent_rigid_ax, delta_ordi, PK, BK, section_ax_name, file_straps_cord):
    
    sz=symbols('z')
    spl_str = PK + percent_rigid_ax * (BK - PK) + delta_ordi
    
    er = 0 #допустимая ошибка в долях
    
    def objective(x):
        return sum(abs(np.dot(A, x) - b))
    
    straps_cord = pd.read_excel(file_straps_cord, sheet_name=name_of_element, index_col=0)
    ordi_center = list(straps_cord.loc[section_ax_name])
    ordi_center = [ordi_center[i] for i in range(len(ordi_center))]
    #ordi_center = [ordi_center[i]/1000 for i in range(len(ordi_center))]
    wb = openpyxl.load_workbook('File_for_result.xlsx')
    worksheet = wb['Лист1']
    start_line = 1
    for i in range(len(ordi_center)):
        print(i)
        straps_cord_absc = straps_cord.iloc[1:, i].tolist()
        for j in range(len(straps_cord_absc)-1, -1, -1):
            if str(straps_cord_absc[j]) == 'nan':
                straps_cord_absc.pop(j)
        if len(straps_cord_absc)>=2:
            straps_cord_absc = [_ for _ in straps_cord_absc]
            for j in range(len(straps_cord_absc)):
                straps_cord_absc[j] = float((straps_cord_absc[j] + delta_ordi - spl_str.subs(sz, ordi_center[i]))/1000)        
            print(straps_cord_absc)

            A = [
                    # Qy
                    [1 for j in range(len(straps_cord_absc))],
                    # Mz
                    [straps_cord_absc[j] for j in range(len(straps_cord_absc))]
                ]

            b = [
                    # Qy
                    sum_Qy[i],
                    # Mz
                    sum_Mz[i]
                ]

            print(A)
            print(b)

            bounds = Bounds([10 for j in range(len(straps_cord_absc))], [np.inf for j in range(len(straps_cord_absc))])
            linear_constraint = LinearConstraint(A, [b[i]*(1 - er) for i in range(len(b))], [b[i]*(1 + er) for i in range(len(b))], keep_feasible=False)
            lsq = minimize(objective, np.zeros([len(straps_cord_absc)]), method='trust-constr', constraints=[linear_constraint], hess=None, bounds=bounds)
            F = list(lsq.x)
            check = np.dot(A, F) - b
            print('Проверка: ', check)
            print('Погрешность %: ', [((np.dot(A[i], F)/b[i])-1)*100 for i in range(len(check))])
            print(F)
            c = worksheet.cell(row=start_line, column = 1)
            c.value = 'Координата в мм:'
            c = worksheet.cell(row=start_line, column = 2)
            c.value = ordi_center[i]
            c = worksheet.cell(row=start_line+1, column = 1)
            c.value = 'Усилия в кгс:'
            start_column = 2
            for j in range(len(F)):
                c = worksheet.cell(row=start_line+1, column = start_column)
                c.value = F[j]
                start_column += 1
            c = worksheet.cell(row=start_line+2, column = 1)
            c.value = 'Погрешность в %:'
            c = worksheet.cell(row=start_line+2, column = 2)
            c.value = str([((np.dot(A[i], F)/b[i])-1)*100 for i in range(len(check))])
            start_line += 4
        else:
            continue
    wb.save("result.xlsx")
    return F, check