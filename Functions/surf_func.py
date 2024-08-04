#!/usr/bin/env python
# coding: utf-8

# In[17]:


import pandas as pd
import numpy as np
from scipy import *
from sympy import *
import math as mp
from scipy.linalg import solve
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit 
import openpyxl
from scipy.interpolate import LinearNDInterpolator


# In[18]:


def MakingSurface(file_straps_cord, section_ax_name, workbook, ord_ax_name):
    #(file_straps_cord - имя файла с кооринатами лямок, workbook - имя файла с распределениями и шаблонами для записи)
    
    sheet_names = pd.ExcelFile(workbook).sheet_names
    dis_p = pd.read_excel(workbook, sheet_name=sheet_names[0], index_col=0)
    
    """Верхняя часть"""

    dis_p_data = [[[], []] for _ in range(int((len(dis_p.columns)-1)/2))]

    absc, ordi, p =[], [], []

    """Запись ординат сечений"""
    ordi_list = dis_p[ord_ax_name].tolist()
    ordi_list = [x for x in ordi_list if str(x) != 'nan']

    """Запись абсцисс"""
    list_dis_p_absc = dis_p.iloc[:, 0:(len(ordi_list)*2):2].values.tolist()
    for i in range(len(list_dis_p_absc[0])):
        for j in range(len(list_dis_p_absc)):
            dis_p_data[i][0].append(list_dis_p_absc[j][i])

    """Запись p"""
    list_dis_p = dis_p.iloc[:, 1:(len(ordi_list)*2):2].values.tolist()
    for i in range(len(list_dis_p[0])):
        for j in range(len(list_dis_p)):
            dis_p_data[i][1].append(list_dis_p[j][i])

    num_points_sec = int(len([x for x in dis_p.iloc[:, 0].values.tolist() if str(x) != 'nan'])/2) #the number of points in the section
            
    for i in range(len(ordi_list)):
        for j in range(num_points_sec):
            absc.append(list_dis_p_absc[j][i])
            ordi.append(ordi_list[i])
            p.append(list_dis_p[j][i])

    X = np.linspace(min(absc), max(absc), num=num_points_sec)
    Z = np.linspace(min(ordi), max(ordi), num=int((max(ordi)-min(ordi))*1000))
    X, Z = np.meshgrid(X, Z)
    # Interpolate the z values on the meshgrid
    interp = LinearNDInterpolator(list(zip(absc, ordi)), p)
    P = interp(X, Z)
    fig = plt.figure(figsize=(10,10)) 
    plt.pcolormesh(X, Z, P, shading='auto')
    plt.plot(absc, ordi, "ok", label="input point")
    plt.legend()
    plt.colorbar()
    plt.axis("equal")
    plt.show()

    straps_cord = pd.read_excel(file_straps_cord, sheet_name=pd.ExcelFile(file_straps_cord).sheet_names[0], index_col=0)

    ordi_centre = list(straps_cord.loc[section_ax_name])
    ordi_centre = [ordi_centre[i]/1000 for i in range(len(ordi_centre))]

    absc_s, ordi_s, p_s = [], [], []

    for i in range(len(ordi_centre)):
        for j in range(len(Z)):
            if round(ordi_centre[i]*1000)==round(Z[j][0]*1000):
                for k in range(num_points_sec):
                    absc_s.append(X[j][k])
                    ordi_s.append(ordi_centre[i])
                    p_s.append(P[j][k])

    # Выбираем нужный лист
    wb = openpyxl.load_workbook(workbook)
    worksheet = wb[sheet_names[1]]

    for i in range(len(ordi_centre)):
        c = worksheet.cell(row=1, column=(2*i)+2)
        c.value = "x=" + str(ordi_centre[i])
        c = worksheet.cell(row=1, column=(2*i+1)+2)
        c.value = "p=" + str(ordi_centre[i])
        for j in range(num_points_sec):
            c = worksheet.cell(row=j+2, column=(2*i)+2)
            c.value = absc_s[i*num_points_sec+j]
            c = worksheet.cell(row=j+2, column=(2*i+1)+2)
            c.value = p_s[i*num_points_sec+j]
    
    for l in range(2, len(sheet_names)):
        ordi_type = pd.read_excel(file_straps_cord, sheet_name=sheet_names[l], index_col=0)
        ordi_type_list = list(ordi_type.loc[section_ax_name])
        ordi_type_list = [ordi_type_list[i]/1000 for i in range(len(ordi_type_list))]
        # Выбираем нужный лист
        worksheet = wb[sheet_names[l]]
        for i in range(len(ordi_centre)):
            for k in range(len((ordi_type_list))):
                if ordi_type_list[k]==ordi_centre[i]:
                    c = worksheet.cell(row=1, column=(2*i)+2)
                    c.value = ordi_type_list[k]
                    c = worksheet.cell(row=1, column=(2*i+1)+2)
                    c.value = ordi_type_list[k]
                    for j in range(num_points_sec):
                        c = worksheet.cell(row=j+2, column=(2*i)+2)
                        c.value = absc_s[i*num_points_sec+j]
                        c = worksheet .cell(row=j+2, column=(2*i+1)+2)
                        c.value = p_s[i*num_points_sec+j]

    """Нижняя часть"""
     
    dis_p_data = [[[], []] for _ in range(int((len(dis_p.columns)-1)/2))]

    absc, ordi, p =[], [], []

    """Запись X"""
    list_dis_p_absc = dis_p.iloc[:, 0:(len(ordi_list)*2):2].values.tolist()
    for i in range(len(list_dis_p_absc[0])):
        for j in range(len(list_dis_p_absc)):
            dis_p_data[i][0].append(list_dis_p_absc[j][i])

    """Запись p"""
    list_dis_p = dis_p.iloc[:, 1:(len(ordi_list)*2):2].values.tolist()
    for i in range(len(list_dis_p[0])):
        for j in range(len(list_dis_p)):
            dis_p_data[i][1].append(list_dis_p[j][i])
            
    for i in range(len(ordi_list)):
        for j in range(num_points_sec, num_points_sec*2):
            absc.append(list_dis_p_absc[j][i])
            ordi.append(ordi_list[i])
            p.append(list_dis_p[j][i])

    X = np.linspace(min(absc), max(absc), num=num_points_sec)
    Z = np.linspace(min(ordi), max(ordi), num=int((max(ordi)-min(ordi))*1000))
    X, Z = np.meshgrid(X, Z)
    # Interpolate the z values on the meshgrid
    interp = LinearNDInterpolator(list(zip(absc, ordi)), p)
    P = interp(X, Z)
    fig = plt.figure(figsize=(10,10)) 
    plt.pcolormesh(X, Z, P, shading='auto')
    plt.plot(absc, ordi, "ok", label="input point")
    plt.legend()
    plt.colorbar()
    plt.axis("equal")
    plt.show()

    absc_s, ordi_s, p_s = [], [], []

    for i in range(len(ordi_centre)):
        for j in range(len(Z)):
            if round(ordi_centre[i]*1000)==round(Z[j][0]*1000):
                for k in range(num_points_sec):
                    absc_s.append(X[j][k])
                    ordi_s.append(ordi_centre[i])
                    p_s.append(P[j][k])
    
    worksheet = wb[sheet_names[1]]
    
    for i in range(len(ordi_centre)):
        for j in range(num_points_sec):
            c = worksheet.cell(row=j+2+num_points_sec, column=(2*i)+2)
            c.value = absc_s[i*num_points_sec+j]
            c = worksheet.cell(row=j+2+num_points_sec, column=(2*i+1)+2)
            c.value = p_s[i*num_points_sec+j]
    
    for l in range(2, len(sheet_names)):
        ordi_type = pd.read_excel(file_straps_cord, sheet_name=sheet_names[l], index_col=0)
        ordi_type_list = list(ordi_type.loc[section_ax_name])
        ordi_type_list = [ordi_type_list[i]/1000 for i in range(len(ordi_type_list))]
        # Выбираем нужный лист
        worksheet = wb[sheet_names[l]]
        for i in range(len(ordi_centre)):
            for k in range(len((ordi_type_list))):
                if ordi_type_list[k]==ordi_centre[i]:
                    for j in range(num_points_sec):
                        c = worksheet.cell(row=j+2+num_points_sec, column=(2*i)+2)
                        c.value = absc_s[i*num_points_sec+j]
                        c = worksheet.cell(row=j+2+num_points_sec, column=(2*i+1)+2)
                        c.value = p_s[i*num_points_sec+j]

    # Сохраняем изменения в файле
    wb.save('Распределение_давлений_по_хордам_с_лямками.xlsx')


# In[ ]:




